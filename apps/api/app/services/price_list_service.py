import os
import uuid
from typing import Optional, Dict, List, Union
from datetime import datetime
import tempfile
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from werkzeug.datastructures import FileStorage
from app.database.models import SupplierPriceList, PriceListItem, Material, Supplier
from app.database.repositories import SupplierPriceListRepository, PriceListItemRepository, MaterialRepository, SupplierRepository

class PriceListService:

    def __init__(self, db: Session):
        self.db = db
        self.price_list_repo = SupplierPriceListRepository(db)
        self.price_list_item_repo = PriceListItemRepository(db)
        self.material_repo = MaterialRepository(db)
        self.supplier_repo = SupplierRepository(db)
        
    def create_template(self, supplier_id: str) -> bytes:
        supplier = self.supplier_repo.get_by_id(supplier_id)
        if not supplier:
            raise ValueError(f"Поставщик с ID {supplier_id} не найден")
        
        template_df = pd.DataFrame(columns=[
            'КодТовараПоставщика',
            'ШтрихКод',
            'Наименование',
            'Артикул',
            'Описание',
            'СтавкаНДС',
            'Цена',
            'Доступность'
        ])
        
        template_df.loc[0] = [
            'ABC123', 
            '4601234567890', 
            'Пример товара', 
            'ART-001', 
            'Описание товара', 
            20, 
            1000.00, 
            10
        ]
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            template_df.to_excel(temp_file.name, index=False, sheet_name='Прайс-лист')
            workbook = openpyxl.load_workbook(temp_file.name)
            worksheet = workbook.active
            worksheet.insert_rows(0, 3)
            worksheet['A1'] = f"Шаблон прайс-листа для поставщика: {supplier.name}"
            worksheet['A2'] = f"Дата создания шаблона: {datetime.now().strftime('%d.%m.%Y')}"
            
            header_row = 4  
            header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
            header_font = Font(bold=True)
            
            for col in range(1, len(template_df.columns) + 1):
                cell = worksheet.cell(row=header_row, column=col)
                cell.fill = header_fill
                cell.font = header_font
                column_letter = get_column_letter(col)
                if col == 3:  
                    worksheet.column_dimensions[column_letter].width = 40
                elif col == 5:  
                    worksheet.column_dimensions[column_letter].width = 50
                else:
                    worksheet.column_dimensions[column_letter].width = 20

            worksheet.insert_rows(header_row + 2, 4)
            notes_row = header_row + 2
            worksheet.merge_cells(f'A{notes_row}:H{notes_row}')
            worksheet['A' + str(notes_row)] = "ИНСТРУКЦИЯ ПО ЗАПОЛНЕНИЮ:"
            worksheet['A' + str(notes_row)].font = Font(bold=True)
            worksheet.merge_cells(f'A{notes_row+1}:H{notes_row+1}')
            worksheet['A' + str(notes_row+1)] = "1. Обязательные поля: 'Наименование', 'Цена'"
            worksheet.merge_cells(f'A{notes_row+2}:H{notes_row+2}')
            worksheet['A' + str(notes_row+2)] = "2. Поле 'СтавкаНДС' заполняется числом (например, 20 для НДС 20%)"
            worksheet.merge_cells(f'A{notes_row+3}:H{notes_row+3}')
            worksheet['A' + str(notes_row+3)] = "3. Поле 'Доступность' указывает количество товара в наличии (0 - нет в наличии)"
            workbook.save(temp_file.name)
            
            with open(temp_file.name, 'rb') as f:
                template_bytes = f.read()
                
        os.unlink(temp_file.name)
        
        return template_bytes
        
    def process_price_list(self, supplier_id: str, date_str: str, file_data: bytes, file_name: str) -> dict:
        temp_file = None
        try:
            supplier = self.supplier_repo.get_by_id(supplier_id)
            if not supplier:
                raise ValueError(f"Поставщик с ID {supplier_id} не найден")

            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            active_price_lists = self.price_list_repo.get_all(supplier_id=supplier_id, is_active=True)
            for pl in active_price_lists:
                self.price_list_repo.deactivate(pl.id)

            temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            temp_file.write(file_data)
            temp_file.close()  
                
            try:
                df = pd.read_excel(temp_file.name, engine='openpyxl')
            except Exception as e:
                raise ValueError(f"Ошибка при чтении файла: {str(e)}")
                
            expected_columns = ['Наименование', 'Цена']
            missing_columns = [col for col in expected_columns if col not in df.columns]
            if missing_columns:
                raise ValueError(f"В файле отсутствуют обязательные столбцы: {', '.join(missing_columns)}")
            

            column_mapping = {
                'КодТовараПоставщика': 'supplier_code',
                'ШтрихКод': 'barcode',
                'Наименование': 'name',
                'Артикул': 'article',
                'Описание': 'description',
                'СтавкаНДС': 'vat_rate',
                'Цена': 'price',
                'Доступность': 'availability'
            }

            rename_dict = {}
            for rus_col, eng_col in column_mapping.items():
                if rus_col in df.columns:
                    rename_dict[rus_col] = eng_col

            if rename_dict:
                df = df.rename(columns=rename_dict)

            df = df.dropna(how='all')
            df = df.dropna(subset=['name', 'price'])

            if 'supplier_code' not in df.columns:
                df['supplier_code'] = None
            if 'barcode' not in df.columns:
                df['barcode'] = None
            if 'article' not in df.columns:
                df['article'] = None
            if 'description' not in df.columns:
                df['description'] = None
            if 'vat_rate' not in df.columns:
                df['vat_rate'] = 20.0
            if 'availability' not in df.columns:
                df['availability'] = 0

            items_data = df.to_dict('records')
            for item in items_data:
                for key, value in list(item.items()):
                    if pd.isna(value):
                        item[key] = None
            price_list_data = {
                'supplier_id': supplier_id,
                'date': date_obj,
                'file_name': file_name,
                'is_active': True
            }
            
            price_list = self.price_list_repo.create(price_list_data, items_data)
            
            self._link_items_to_materials(price_list.id)
            
            return {
                'price_list_id': str(price_list.id),
                'supplier_id': str(supplier_id),
                'date': date_str,
                'processed_items': len(items_data),
                'file_name': file_name
            }
                
        except Exception as e:
            raise ValueError(f"Ошибка при обработке прайс-листа: {str(e)}")
        finally:
            if temp_file and os.path.exists(temp_file.name):
                try:
                    os.unlink(temp_file.name)
                except (PermissionError, OSError):
                    pass
    
    def deactivate_price_list(self, price_list_id: str) -> bool:
        return self.price_list_repo.deactivate(price_list_id)
    
    def activate_price_list(self, price_list_id: str) -> bool:
        price_list = self.price_list_repo.get_by_id(price_list_id)
        if not price_list:
            raise ValueError(f"Прайс-лист с ID {price_list_id} не найден")
        
        price_list.is_active = True
        price_list.updated_at = datetime.now()
        self.db.commit()
        return True
    
    def get_price_list_details(self, price_list_id: str) -> dict:
        price_list = self.price_list_repo.get_by_id(price_list_id)
        if not price_list:
            raise ValueError(f"Прайс-лист с ID {price_list_id} не найден")
        
        return {
            'id': str(price_list.id),
            'supplier_id': str(price_list.supplier_id),
            'supplier_name': price_list.supplier.name,
            'date': price_list.date.isoformat(),
            'file_name': price_list.file_name,
            'is_active': price_list.is_active,
            'created_at': price_list.created_at.isoformat(),
            'updated_at': price_list.updated_at.isoformat()
        }
    
    def get_price_list_items(self, price_list_id: str, search: str = None, skip: int = 0, limit: int = 100) -> list:
        items = self.price_list_item_repo.get_by_price_list(price_list_id, skip, limit, search)
        
        result = []
        for item in items:
            item_dict = {
                'id': str(item.id),
                'price_list_id': str(item.price_list_id),
                'supplier_code': item.supplier_code,
                'barcode': item.barcode,
                'name': item.name,
                'article': item.article,
                'description': item.description,
                'vat_rate': item.vat_rate,
                'price': item.price,
                'availability': item.availability,
                'material_id': str(item.material_id) if item.material_id else None,
                'created_at': item.created_at.isoformat(),
                'updated_at': item.updated_at.isoformat()
            }
            
            if item.material_id:
                material = self.material_repo.get_by_id(item.material_id)
                if material:
                    item_dict['material'] = {
                        'id': str(material.id),
                        'name': material.name,
                        'unit': material.unit,
                        'description': material.description
                    }
            
            result.append(item_dict)
        
        return result

    def get_current_price_for_material(self, material_id: str, supplier_id: str) -> Optional[dict]:
        try:
            item = self.price_list_item_repo.get_current_price_for_material(material_id, supplier_id)
            if not item:
                return None
            price_list = self.price_list_repo.get_by_id(item.price_list_id)
            return {
                'id': str(item.id),
                'price_list_id': str(item.price_list_id),
                'price_list_date': price_list.date.isoformat(),
                'supplier_id': str(price_list.supplier_id),
                'material_id': str(item.material_id),
                'price': item.price,
                'vat_rate': item.vat_rate,
                'availability': item.availability,
                'is_active_price_list': price_list.is_active
            }
        except Exception as e:
            print(f"Ошибка при получении актуальной цены: {str(e)}")
            return None
        
    def _link_items_to_materials(self, price_list_id: str) -> int:
        items = self.price_list_item_repo.get_by_price_list(price_list_id)

        if not items:
            return 0

        materials = self.material_repo.get_all()
        materials_by_name = {}
        materials_by_article = {}
        materials_by_barcode = {}
        materials_by_supplier_code = {}
        
        for m in materials:
            if m.name:
                materials_by_name[m.name.lower()] = m

            if m.description:
                article_match = None
                barcode_match = None
                supplier_code_match = None
                for line in m.description.split('\n'):
                    line = line.strip()
                    if 'артикул:' in line.lower():
                        parts = line.split(':', 1)
                        if len(parts) > 1 and parts[1].strip():
                            article_match = parts[1].strip()
                    elif 'штрихкод:' in line.lower() or 'ean:' in line.lower() or 'штрих-код:' in line.lower():
                        parts = line.split(':', 1)
                        if len(parts) > 1 and parts[1].strip():
                            barcode_match = parts[1].strip()
                    elif 'код поставщика:' in line.lower() or 'код товара:' in line.lower():
                        parts = line.split(':', 1)
                        if len(parts) > 1 and parts[1].strip():
                            supplier_code_match = parts[1].strip()
                if article_match:
                    materials_by_article[article_match.lower()] = m
                if barcode_match:
                    materials_by_barcode[barcode_match] = m
                if supplier_code_match:
                    materials_by_supplier_code[supplier_code_match.lower()] = m
        
        linked_count = 0
        for item in items:
            if item.material_id:
                continue
            if item.barcode and item.barcode in materials_by_barcode:
                item.material_id = materials_by_barcode[item.barcode].id
                linked_count += 1
                continue
            if item.article and item.article.lower() in materials_by_article:
                item.material_id = materials_by_article[item.article.lower()].id
                linked_count += 1
                continue
            if item.supplier_code and item.supplier_code.lower() in materials_by_supplier_code:
                item.material_id = materials_by_supplier_code[item.supplier_code.lower()].id
                linked_count += 1
                continue
            if item.name and item.name.lower() in materials_by_name:
                item.material_id = materials_by_name[item.name.lower()].id
                linked_count += 1
                continue
            if item.name:
                item_name_lower = item.name.lower()
                best_match = None
                best_match_score = 0
                for material_name, material in materials_by_name.items():
                    if material_name in item_name_lower or item_name_lower in material_name:
                        common_length = min(len(material_name), len(item_name_lower))
                        if common_length > best_match_score:
                            best_match = material
                            best_match_score = common_length
                if best_match and best_match_score > 5: 
                    item.material_id = best_match.id
                    linked_count += 1
        if linked_count > 0:
            self.db.commit()
        return linked_count
    
    def link_item_to_material(self, item_id: str, material_id: str) -> bool:
        return self.price_list_item_repo.link_to_material(item_id, material_id)
    
    def update_price_list_item(self, item_id: str, item_data: dict) -> dict:
        item = self.price_list_item_repo.update(item_id, item_data)
        if not item:
            raise ValueError(f"Позиция прайс-листа с ID {item_id} не найдена")
        return {
            'id': str(item.id),
            'price_list_id': str(item.price_list_id),
            'supplier_code': item.supplier_code,
            'barcode': item.barcode,
            'name': item.name,
            'article': item.article,
            'description': item.description,
            'vat_rate': item.vat_rate,
            'price': item.price,
            'availability': item.availability,
            'material_id': str(item.material_id) if item.material_id else None,
            'created_at': item.created_at.isoformat(),
            'updated_at': item.updated_at.isoformat()
        }